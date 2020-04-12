import openpyxl
import  os
os.chdir('excel')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))
sheet = workbook['Sheet1']
print(type(sheet))

cellA1 = sheet['A1'].value
print(cellA1)
sheet.cell(row=1, column=2)
print(sheet.cell(row=1, column=2))

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
