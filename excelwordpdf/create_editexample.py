import openpyxl
import os

os.chdir('excel')
wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb['Sheet']

sheet['A1'].value
# sheet['A1'].value == None

sheet['A1'] = 42
sheet['A2'] = 'Hello'

# wb.save('exampleOpenEdit.xlsx')

sheet2 = wb.create_sheet()
sheet2.title = 'My new Sheet name'
print(wb.sheetnames)
wb.save('exampleOpenEdit2.xlsx')

wb.create_sheet(index=0, title='My other sheet')
wb.save('exampleOpenEdit3.xlsx')
